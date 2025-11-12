# import required python modules
import pandas as pd
import openpyxl

class ParamValues:
    def __init__(self, excel_file):
        self.excel_file = excel_file

    def get_param_list(self):
        fc_dict = {}
        # Read Excel Data
        excel_data = pd.read_excel(self.excel_file, sheet_name=None, keep_default_na=False)
        
        # Data Prep Clean
        fc_to_create_buffer_zone = list(excel_data["1_DataPreparation"].loc[excel_data["1_DataPreparation"]["Function Name"] == "FC to create buffer zone", "FeatureClass"])
        fc_dict['buffer_points_25K'] = fc_to_create_buffer_zone
        not_include_fields = list(excel_data["1_DataPreparation"].loc[excel_data["1_DataPreparation"]["Feature Usage Notes"] == "Not include fields", "Field"])
        fc_dict['not_include_fields'] = not_include_fields
        feature_to_split = list(excel_data["1_DataPreparation"].loc[excel_data["1_DataPreparation"]["Feature Usage Notes"] == "Feature_to_split", "FeatureClass"])
        fc_dict['feature_to_split'] = feature_to_split
        bau_field_fc = list(excel_data["1_DataPreparation"].loc[excel_data["1_DataPreparation"]["Feature Usage Notes"] == "BAU Field FC", "FeatureClass"])
        fc_dict['bau_field_fc'] = bau_field_fc

        # Transportation
        trans_build_up_buildings = list(excel_data["2_Transportation"].loc[excel_data["2_Transportation"]["Feature Usage Notes"] == "Compare Features", "FeatureClass"])
        fc_dict['trans_build_up_buildings'] = trans_build_up_buildings
        topology_features = list(excel_data["2_Transportation"].loc[excel_data["2_Transportation"]["Feature Usage Notes"] == "Topology Features", "FeatureClass"])
        fc_dict['topology_features'] = topology_features
        collapse_sql = list(excel_data["2_Transportation"].loc[excel_data["2_Transportation"]["Function Name"] == "Collapse Road Detail & Replace", "Query"])
        fc_dict['collapse_sql'] = collapse_sql
        railway_sql = list(excel_data["2_Transportation"].loc[excel_data["2_Transportation"]["Function Name"] == "Convert close single track to double", "Expression"])
        fc_dict['railway_sql'] = railway_sql
        fcs_trim_extent_trans = list(excel_data["2_Transportation"].loc[excel_data["2_Transportation"]["Function Name"] == "DATA_TRIM_EXTEND", "FeatureClass"])
        fc_dict['fcs_trim_extent_trans'] = fcs_trim_extent_trans
        sql = list(excel_data["2_Transportation"].loc[excel_data["2_Transportation"]["Function Name"] == "Feature to Point", "Query"])
        fc_dict['sql'] = sql

        # Hydrography
        fcs_trim_extent_hyd = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Function Name"] == "DATA_TRIM_EXTEND", "FeatureClass"])
        fc_dict['fcs_trim_extent_hyd'] = fcs_trim_extent_hyd
        hydro_prep_fc_list = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Feature Usage Notes"] == "Hydro Prep", "FeatureClass"])
        fc_dict['hydro_prep_fc_list'] = hydro_prep_fc_list
        hydro_input_polygon_fc = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Feature Usage Notes"] == "Input polygon Feature Class", "FeatureClass"])
        fc_dict['hydro_input_polygon_fc'] = hydro_input_polygon_fc
        hydro_center_line_fc = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Feature Usage Notes"] == "Centerlines Feature Class", "FeatureClass"])
        fc_dict['hydro_center_line_fc'] = hydro_center_line_fc
        hydro_replace_fc = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Feature Usage Notes"] == "Replace Feature Class", "FeatureClass"])
        fc_dict['hydro_replace_fc'] = hydro_replace_fc
        hydro_remove_near_poly_list = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Function Name"] == "Hydro Remove Near Polygons", "FeatureClass"])
        fc_dict['hydro_remove_near_poly_list'] = hydro_remove_near_poly_list
        hydro_enlarge_poly_list = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Function Name"] == "Hydro Enlarge Polygons", "FeatureClass"])
        fc_dict['hydro_enlarge_poly_list'] = hydro_enlarge_poly_list
        hydro_remove_small_poly_list = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Function Name"] == "Hydro Remove Small Polygons by Converting", "FeatureClass"])
        fc_dict['hydro_remove_small_poly_list'] = hydro_remove_small_poly_list
        hydro_erase_poly_list = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Function Name"] == "Hydro Erase Polygons", "FeatureClass"])
        fc_dict['hydro_erase_poly_list'] = hydro_erase_poly_list
        hydro_small_line_fc_list = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Feature Usage Notes"] == "Feature 2 Point Line", "FeatureClass"])
        fc_dict['hydro_small_line_fc_list'] = hydro_small_line_fc_list
        hydro_small_point_fc_list = list(excel_data["3_Hydrography"].loc[excel_data["3_Hydrography"]["Feature Usage Notes"] == "Feature 2 Point Point", "FeatureClass"])
        fc_dict['hydro_small_point_fc_list'] = hydro_small_point_fc_list

        # Built-up Generalization
        small_bldg_2_point_a = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Feature Usage Notes"] == "Polygon", "FeatureClass"])
        fc_dict['small_bldg_2_point_a'] = small_bldg_2_point_a
        small_bldg_2_point_p = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Feature Usage Notes"] == "Point", "FeatureClass"])
        fc_dict['small_bldg_2_point_p'] = small_bldg_2_point_p
        features_in_cemetery = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Function Name"] == "Delete Buildings in Cemetery", "FeatureClass"])
        fc_dict['features_in_cemetery'] = features_in_cemetery
        enlarge_barrier_features = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Function Name"] == "Enlarge Builtup Features (Cemetery)", "FeatureClass"])
        fc_dict['enlarge_barrier_features'] = enlarge_barrier_features
        delete_small_bldgs = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Function Name"] == "Delete Small Buildings", "FeatureClass"])
        fc_dict['delete_small_bldgs'] = delete_small_bldgs
        enlarge_building_features = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Function Name"] == "Enlarge Small Buildings", "FeatureClass"])
        fc_dict['enlarge_building_features'] = enlarge_building_features
        delineate_building_layers = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Feature Usage Notes"] == "Input Building Layers", "FeatureClass"])
        fc_dict['delineate_building_layers'] = delineate_building_layers
        delineate_edge_features = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Feature Usage Notes"] == "Edge Features", "FeatureClass"])
        fc_dict['delineate_edge_features'] = delineate_edge_features
        delete_small_features = list(excel_data["4_Built Environment"].loc[excel_data["4_Built Environment"]["Function Name"] == "Delete Small Features (Swimming)", "FeatureClass"])
        fc_dict['delete_small_features'] = delete_small_features

        # Vegetation
        veg_lyrs_list = list(excel_data["7_Vegetation"].loc[excel_data["7_Vegetation"]["Function Name"] == "Vegetation", "FeatureClass"])
        fc_dict['veg_lyrs_list'] = veg_lyrs_list
        veg_field_values = list(excel_data["7_Vegetation"].loc[excel_data["7_Vegetation"]["Rule Type"] == "Calculate", "Value"])
        fc_dict['veg_field_values'] = veg_field_values

        # Utility
        utility_area_features = list(excel_data["5_Utility"].loc[excel_data["5_Utility"]["Feature Usage Notes"] == "Polygon", "FeatureClass"])
        fc_dict['utility_area_features'] = utility_area_features
        utility_point_features = list(excel_data["5_Utility"].loc[excel_data["5_Utility"]["Feature Usage Notes"] == "Point", "FeatureClass"])
        fc_dict['utility_point_features'] = utility_point_features
        utility_compare_features = list(excel_data["5_Utility"].loc[excel_data["5_Utility"]["Feature Usage Notes"] == "Compare Features", "FeatureClass"])
        fc_dict['utility_compare_features'] = utility_compare_features

        # Hypsography
        hypso_compare_features = list(excel_data["6_Hypsography"].loc[excel_data["6_Hypsography"]["Function Name"] == "Erase Vegetation Hypso", "FeatureClass"])
        fc_dict['hypso_compare_features'] = hypso_compare_features

        # Detect conflict
        Structure2Structure = list(excel_data["10_DetectConflicts"].loc[excel_data["10_DetectConflicts"]["Function Name"] == "Structure2Structure", "FeatureClass"])
        fc_dict['Structure2Structure'] = Structure2Structure
        Structure2Lines = list(excel_data["10_DetectConflicts"].loc[excel_data["10_DetectConflicts"]["Function Name"] == "Structure2Lines", "FeatureClass"])
        fc_dict['Structure2Lines'] = Structure2Lines
        Lines2Lines = list(excel_data["10_DetectConflicts"].loc[excel_data["10_DetectConflicts"]["Function Name"] == "Lines2Lines", "FeatureClass"])
        fc_dict['Lines2Lines'] = Lines2Lines
        G1_Poly2Poly = list(excel_data["10_DetectConflicts"].loc[excel_data["10_DetectConflicts"]["Function Name"] == "G1_Poly2Poly", "FeatureClass"])
        fc_dict['G1_Poly2Poly'] = G1_Poly2Poly
        G2_Poly2Poly = list(excel_data["10_DetectConflicts"].loc[excel_data["10_DetectConflicts"]["Function Name"] == "G2_Poly2Poly", "FeatureClass"])
        fc_dict['G2_Poly2Poly'] = G2_Poly2Poly
        G3_Poly2Poly = list(excel_data["10_DetectConflicts"].loc[excel_data["10_DetectConflicts"]["Function Name"] == "G3_Poly2Poly", "FeatureClass"])
        fc_dict['G3_Poly2Poly'] = G3_Poly2Poly

        # Resolve conflicts buildings
        built_up_area_fcs = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Function Name"] == "Hide Buildings Under Generalized and Town Built-Up Area", "FeatureClass"])
        fc_dict['build_up_area_fcs'] = built_up_area_fcs
        input_building_layers = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "Input Building Layers", "FeatureClass"])
        fc_dict['input_building_layers'] = input_building_layers
        input_barrier_layers = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "Input Barrier Layers", "FeatureClass"])
        fc_dict['input_barrier_layers'] = input_barrier_layers
        g5_input_points = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "G5 Input Points", "FeatureClass"])
        fc_dict['g5_input_points'] = g5_input_points
        g7_input_points = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "G7 Input Points", "FeatureClass"])
        fc_dict['g7_input_points'] = g7_input_points
        g1_align_features = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "G1 Align Features", "FeatureClass"])
        fc_dict['g1_align_features'] = g1_align_features
        g4_align_features = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "G4 Align Features", "FeatureClass"])
        fc_dict['g4_align_features'] = g4_align_features
        g5_align_features = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "G5 Align Features", "FeatureClass"])
        fc_dict['g5_align_features'] = g5_align_features
        g6_align_features = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "G6 Align Features", "FeatureClass"])
        fc_dict['g6_align_features'] = g6_align_features
        g7_align_features = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "G7 Align Features", "FeatureClass"])
        fc_dict['g7_align_features'] = g7_align_features
        input_primary = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "Convert Polygons", "FeatureClass"])
        fc_dict['input_primary'] = input_primary
        input_secondary = list(excel_data["9b_ResolveConflictsBuildings"].loc[excel_data["9b_ResolveConflictsBuildings"]["Feature Usage Notes"] == "Compare features", "FeatureClass"])
        fc_dict['input_secondary'] = input_secondary

        # Resolve conflicts lines
        input_line_layers = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Feature Usage Notes"] == "Input Line Layers", "FeatureClass"])
        fc_dict['input_line_layers'] = input_line_layers
        edge_features = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Feature Usage Notes"] == "Edge Feature", "FeatureClass"])
        fc_dict['edge_features'] = edge_features
        embank_list = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Feature Usage Notes"] == "Embankment Features", "FeatureClass"])
        fc_dict['embank_list'] = embank_list
        compare_fcs_embank = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Feature Usage Notes"] == "Compare Features", "FeatureClass"])
        fc_dict['compare_fcs_embank'] = compare_fcs_embank
        road_query = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Feature Usage Notes"] == "road_query", "Expression"])
        fc_dict['road_query'] = road_query
        bridge_query = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Feature Usage Notes"] == "bridge_query", "Expression"])
        fc_dict['bridge_query'] = bridge_query
        footprint_fcs = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Function Name"] == "Snap Points to Shifted Lines", "FeatureClass"])
        fc_dict['footprint_fcs'] = footprint_fcs
        resolve_line_compare = list(excel_data["9a_ResolveConflictsLines"].loc[excel_data["9a_ResolveConflictsLines"]["Function Name"] == "Resolve Conflicts for Lakes and Ponds", "FeatureClass"])
        fc_dict['resolve_line_compare'] = resolve_line_compare

        # Apply carto symbology
        attribution_fc_list = list(excel_data["8_ApplyCartoSymbology"].loc[excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Apply Attribution", "FeatureClass"])
        fc_dict['attribution_fc_list'] = attribution_fc_list
        express_list = list(excel_data["8_ApplyCartoSymbology"].loc[excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Apply Attribution", "Expression"])
        fc_dict['express_list'] = express_list
        query_list = list(excel_data["8_ApplyCartoSymbology"].loc[excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Apply Attribution", "Query"])
        fc_dict['query_list'] = query_list
        field_list = list(excel_data["8_ApplyCartoSymbology"].loc[excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Apply Attribution", "Field"])
        fc_dict['field_list'] = field_list
        intersecting_fc_list = list(excel_data["8_ApplyCartoSymbology"].loc[excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Split Embankments and Cuttings", "FeatureClass"])
        fc_dict['intersecting_fc_list'] = intersecting_fc_list
        prep_line_resolve_fcs_list = list(excel_data["8_ApplyCartoSymbology"].loc[excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Prep For Line Resolve", "FeatureClass"])
        fc_dict['prep_line_resolve_fcs_list'] = prep_line_resolve_fcs_list
        apply_symbology_layers_list = list(excel_data["8_ApplyCartoSymbology"].loc[excel_data["8_ApplyCartoSymbology"]["Function Name"] == "Calculate VST on Workspace", "FeatureClass"])
        fc_dict['apply_symbology_layers_list'] = apply_symbology_layers_list
        return fc_dict
    
    def get_param_vals(self):
        val_dict = {}
        # Read Excel Data
        rule_book = openpyxl.load_workbook(self.excel_file)
        # Data Prep Clean
        wb = rule_book['1_DataPreparation']
        dataset_name = wb.cell(row=8, column=10).value
        val_dict['dataset_name'] = dataset_name
        buffer_distance_point = wb.cell(row=2, column=10).value
        val_dict['buffer_distance_point'] = buffer_distance_point
        extend_val = wb.cell(row=3, column=10).value
        val_dict['extend_val'] = extend_val
        trim_val = wb.cell(row=4, column=10).value
        val_dict['trim_val'] = trim_val
        buffer_distance = wb.cell(row=5, column=10).value
        val_dict['buffer_distance'] = buffer_distance
        feature_count = wb.cell(row=6, column=10).value
        val_dict['feature_count'] = feature_count
        vertex_limit = wb.cell(row=7, column=10).value
        val_dict['vertex_limit'] = vertex_limit
        map_name_apply_carto = wb.cell(row=9, column=10).value
        val_dict['map_name_apply_carto'] = map_name_apply_carto
        map_name_resolve_lines = wb.cell(row=10, column=10).value
        val_dict['map_name_resolve_lines'] = map_name_resolve_lines
        map_name_resolve_polygons = wb.cell(row=11, column=10).value
        val_dict['map_name_resolve_polygons'] = map_name_resolve_polygons
        map_name_detect = wb.cell(row=12, column=10).value
        val_dict['map_name_detect'] = map_name_detect

        # Transportation
        wb = rule_book['2_Transportation']
        collapse_size = wb.cell(row=2, column=10).value
        val_dict['collapse_size'] = collapse_size
        trans_common_express = wb.cell(row=4, column=2).value
        val_dict['trans_common_express'] = trans_common_express
        seg_length = wb.cell(row=10, column=10).value
        val_dict['seg_length'] = seg_length
        group_sql_rd1 = wb.cell(row=7, column=2).value
        val_dict['group_sql_rd1'] = group_sql_rd1
        group_sql_rd2 = wb.cell(row=8, column=2).value
        val_dict['group_sql_rd2'] = group_sql_rd2
        group_sql_track = wb.cell(row=9, column=2).value
        val_dict['group_sql_track'] = group_sql_track
        minimum_length_min = wb.cell(row=65, column=10).value
        val_dict['minimum_length_min'] = minimum_length_min
        minimum_length_max = wb.cell(row=64, column=10).value
        val_dict['minimum_length_max'] = minimum_length_max
        simple_tolerance = wb.cell(row=4, column=10).value
        val_dict['simple_tolerance'] = simple_tolerance
        smooth_tolerance = wb.cell(row=5, column=10).value
        val_dict['smooth_tolerance'] = smooth_tolerance
        Merge_Field = wb.cell(row=107, column=5).value
        val_dict['Merge_Field'] = Merge_Field
        min_size = wb.cell(row=101, column=10).value
        val_dict['min_size'] = min_size
        minimum_length = wb.cell(row=105, column=10).value
        val_dict['minimum_length'] = minimum_length
        minimum_width = wb.cell(row=106, column=10).value
        val_dict['minimum_width'] = minimum_width
        merge_distance = wb.cell(row=107, column=10).value
        val_dict['merge_distance'] = merge_distance
        additional_criteria_trans = wb.cell(row=111, column=10).value
        val_dict['additional_criteria_trans'] = additional_criteria_trans
        trans_generalized_operation = wb.cell(row=68, column=10).value
        val_dict['trans_generalized_operation'] = trans_generalized_operation
        trans_delete_input = wb.cell(row=102, column=10).value
        val_dict['trans_delete_input'] = trans_delete_input
        trans_create_one_point= wb.cell(row=104, column=10).value
        val_dict['trans_create_one_point'] = trans_create_one_point
        trans_update_val = wb.cell(row=108, column=10).value
        val_dict['trans_update_val'] = trans_update_val
        trans_changed_road_type = wb.cell(row=109, column=10).value
        val_dict['trans_changed_road_type'] = trans_changed_road_type
        trans_unique_field= wb.cell(row=104, column=5).value
        val_dict['trans_unique_field'] = trans_unique_field

        # Built-up Generalization
        wb = rule_book['4_Built Environment']
        min_size_bldg1 = wb.cell(row=2, column=10).value
        val_dict['min_size_bldg1'] = min_size_bldg1
        sql_bldg = wb.cell(row=5, column=2).value
        val_dict['sql_bldg'] = sql_bldg
        min_size_bldg2 = wb.cell(row=77, column=10).value
        val_dict['min_size_bldg2'] = min_size_bldg2
        sql_bldg1 = wb.cell(row=15, column=2).value
        val_dict['sql_bldg1'] = sql_bldg1
        enlarge_min_size = wb.cell(row=154, column=10).value
        val_dict['enlarge_min_size'] = enlarge_min_size
        enlarge_val = wb.cell(row=155, column=10).value
        val_dict['enlarge_val'] = enlarge_val
        del_min_area = wb.cell(row=159, column=10).value
        val_dict['del_min_area'] = del_min_area
        enlarge_bldg_min_width = wb.cell(row=230, column=10).value
        val_dict['enlarge_bldg_min_width'] = enlarge_bldg_min_width
        enlarge_bldg_min_length = wb.cell(row=231, column=10).value
        val_dict['enlarge_bldg_min_length'] = enlarge_bldg_min_length
        additional_criteria = wb.cell(row=461, column=2).value
        val_dict['additional_criteria'] = additional_criteria
        simpl_bldg_distance = wb.cell(row=232, column=10).value
        val_dict['simpl_bldg_distance'] = simpl_bldg_distance
        delineate_ref_scale = wb.cell(row=285, column=10).value
        val_dict['delineate_ref_scale'] = delineate_ref_scale
        delineate_min_bldg_count = wb.cell(row=286, column=10).value
        val_dict['delineate_min_bldg_count'] = delineate_min_bldg_count
        delineate_min_detail_size = wb.cell(row=287, column=10).value
        val_dict['delineate_min_detail_size'] = delineate_min_detail_size
        delineate_grp_dist = wb.cell(row=288, column=10).value
        val_dict['delineate_grp_dist'] = delineate_grp_dist
        del_small_recreation_fc_min_size = wb.cell(row=393, column=10).value
        val_dict['del_small_recreation_fc_min_size'] = del_small_recreation_fc_min_size
        del_small_recreation_express= wb.cell(row=46, column=2).value
        val_dict['del_small_recreation_express'] = del_small_recreation_express
        erase_sql= wb.cell(row=49, column=2).value
        val_dict['erase_sql'] = erase_sql
        build_delete_input = wb.cell(row=3, column=10).value
        val_dict['build_delete_input'] = build_delete_input
        build_create_one_point= wb.cell(row=4, column=10).value
        val_dict['build_create_one_point'] = build_create_one_point
        build_unique_field= wb.cell(row=2, column=6).value
        val_dict['build_unique_field'] = build_unique_field
        build_update_val = wb.cell(row=233, column=10).value
        val_dict['simplification_tolerance'] = build_update_val

        # Hydrography
        wb = rule_book['3_Hydrography']
        remove_short_line_line_length= wb.cell(row=2, column=10).value
        val_dict['remove_short_line_line_length'] = remove_short_line_line_length
        hydro_np_polygon_width= wb.cell(row=8, column=10).value
        val_dict['hydro_np_polygon_width'] = hydro_np_polygon_width
        hydro_np_polygon_percentage= wb.cell(row=9, column=10).value
        val_dict['hydro_np_polygon_percentage'] = hydro_np_polygon_percentage
        hydro_smooth_tolerance= wb.cell(row=46, column=10).value
        val_dict['hydro_smooth_tolerance'] = hydro_smooth_tolerance
        hydro_simple_tolerance= wb.cell(row=47, column=10).value
        val_dict['hydro_simple_tolerance'] = hydro_simple_tolerance
        hydro_remove_small_poly_exp = wb.cell(row=17, column=2).value
        val_dict['hydro_remove_small_poly_exp'] = hydro_remove_small_poly_exp
        hydro_remove_small_poly_mim_area = wb.cell(row=48, column=10).value
        val_dict['hydro_remove_small_poly_mim_area'] = hydro_remove_small_poly_mim_area
        hydro_enlarge_poly_mim_size = wb.cell(row=49, column=10).value
        val_dict['hydro_enlarge_poly_mim_size'] = hydro_enlarge_poly_mim_size
        hydro_enlarge_poly_buffer_dist = wb.cell(row=50, column=10).value
        val_dict['hydro_enlarge_poly_buffer_dist'] = hydro_enlarge_poly_buffer_dist
        hydro_remove_near_poly_delete_size = wb.cell(row=51, column=10).value
        val_dict['hydro_remove_near_poly_delete_size'] = hydro_remove_near_poly_delete_size
        hydro_remove_near_poly_min_size = wb.cell(row=52, column=10).value
        val_dict['hydro_remove_near_poly_min_size'] = hydro_remove_near_poly_min_size
        hydro_remove_near_poly_dist = wb.cell(row=53, column=10).value
        val_dict['hydro_remove_near_poly_dist'] = hydro_remove_near_poly_dist
        hydro_remove_near_poly_sql = wb.cell(row=22, column=3).value
        val_dict['hydro_remove_near_poly_sql'] = hydro_remove_near_poly_sql
        hydro_enlarge_poly_sql = wb.cell(row=60, column=2).value
        val_dict['hydro_enlarge_poly_sql'] = hydro_enlarge_poly_sql
        hydro_enlarge_untouch_poly_buffer_dist = wb.cell(row=60, column=10).value
        val_dict['hydro_enlarge_untouch_poly_buffer_dist'] = hydro_enlarge_untouch_poly_buffer_dist
        hydro_trim_between_polygon_min_area = wb.cell(row=68, column=10).value
        val_dict['hydro_trim_between_polygon_min_area'] = hydro_trim_between_polygon_min_area
        hydro_trim_between_polygon_distance = wb.cell(row=69, column=10).value
        val_dict['hydro_trim_between_polygon_distance'] = hydro_trim_between_polygon_distance
        hydro_remove_small_min_size = wb.cell(row=70, column=10).value
        val_dict['hydro_remove_small_min_size'] = hydro_remove_small_min_size
        hydro_remove_small_sql = wb.cell(row=70, column=2).value
        val_dict['hydro_remove_small_sql'] = hydro_remove_small_sql
        hydro_erase_poly_max_gap_area = wb.cell(row=78, column=10).value
        val_dict['hydro_erase_poly_max_gap_area'] = hydro_erase_poly_max_gap_area
        hydro_convert_ungr_river_min_length = wb.cell(row=86, column=10).value
        val_dict['hydro_convert_ungr_river_min_length'] = hydro_convert_ungr_river_min_length
        increase_hydro_line_min_length = wb.cell(row=87, column=10).value
        val_dict['increase_hydro_line_min_length'] = increase_hydro_line_min_length
        remove_close_parallel_per_min = wb.cell(row=88, column=10).value
        val_dict['remove_close_parallel_per_min'] = remove_close_parallel_per_min
        remove_close_parallel_per_max = wb.cell(row=89, column=10).value
        val_dict['remove_close_parallel_per_max'] = remove_close_parallel_per_max
        remove_close_dist = wb.cell(row=90, column=10).value
        val_dict['remove_close_dist'] = remove_close_dist
        remove_close_tolerance = wb.cell(row=91, column=10).value
        val_dict['remove_close_tolerance'] = remove_close_tolerance
        hydro_line_dangle_min_length = wb.cell(row=92, column=10).value
        val_dict['hydro_line_dangle_min_length'] = hydro_line_dangle_min_length
        hydro_small_fc_min_length = wb.cell(row=93, column=10).value
        val_dict['hydro_small_fc_min_length'] = hydro_small_fc_min_length
        hydro_generalized_operation = wb.cell(row=103, column=10).value
        val_dict['hydro_generalized_operation'] = hydro_generalized_operation
        hydro_delete_input = wb.cell(row=94, column=10).value
        val_dict['hydro_delete_input'] = hydro_delete_input
        hydro_create_one_point= wb.cell(row=95, column=10).value
        val_dict['hydro_create_one_point'] = hydro_create_one_point
        hydro_trim_update_val = wb.cell(row=96, column=10).value
        val_dict['hydro_trim_update_val'] = hydro_trim_update_val
        hydro_unique_field= wb.cell(row=94, column=6).value
        val_dict['hydro_unique_field'] = hydro_unique_field


        # Vegetation
        wb = rule_book['7_Vegetation']
        vegetation_min_area = wb.cell(row=2, column=10).value
        val_dict['vegetation_min_area'] = vegetation_min_area
        vegetation_eliminate_area = wb.cell(row=3, column=10).value
        val_dict['vegetation_eliminate_area'] = vegetation_eliminate_area

        # Utility
        wb = rule_book['5_Utility']
        utility_min_size_sewerage = wb.cell(row=2, column=10).value
        val_dict['utility_min_size_sewerage'] = utility_min_size_sewerage
        utility_min_size_building = wb.cell(row=3, column=10).value
        val_dict['utility_min_size_building'] = utility_min_size_building
        utility_min_size= wb.cell(row=4, column=10).value
        val_dict['utility_min_size'] = utility_min_size
        utility_beffer_dist= wb.cell(row=5, column=10).value
        val_dict['utility_beffer_dist'] = utility_beffer_dist
        utility_dist = wb.cell(row=6, column=10).value
        val_dict['utility_dist'] = utility_dist
        utility_dist_shorter = wb.cell(row=7, column=10).value
        val_dict['utility_dist_shorter'] = utility_dist_shorter
        utility_addi_criteria_sewerage = wb.cell(row=8, column=10).value
        val_dict['utility_addi_criteria_sewerage'] = utility_addi_criteria_sewerage
        utility_addi_criteria= wb.cell(row=9, column=10).value
        val_dict['utility_addi_criteria'] = utility_addi_criteria
        utility_merge_field = wb.cell(row=12, column=5).value
        val_dict['utility_merge_field'] = utility_merge_field
        utility_delete_input = wb.cell(row=14, column=10).value
        val_dict['utility_delete_input'] = utility_delete_input
        utility_create_one_point= wb.cell(row=15, column=10).value
        val_dict['utility_create_one_point'] = utility_create_one_point
        utility_unique_field= wb.cell(row=2, column=6).value
        val_dict['utility_unique_field'] = utility_unique_field
        utility_update_val= wb.cell(row=16, column=10).value
        val_dict['utility_update_val'] = utility_update_val
        
        # Hypsography
        wb = rule_book['6_Hypsography']
        hypso_dissolved_field = wb.cell(row=12, column=5).value
        val_dict['hypso_dissolved_field'] = hypso_dissolved_field
        hypso_dist = wb.cell(row=2, column=10).value
        val_dict['hypso_dist'] = hypso_dist
        hypso_parallel_per = wb.cell(row=3, column=10).value
        val_dict['hypso_parallel_per'] = hypso_parallel_per
        hypso_min_length = wb.cell(row=4, column=10).value
        val_dict['hypso_min_length'] = hypso_min_length
        hypso_smoothing_tolerance = wb.cell(row=5, column=10).value
        val_dict['hypso_smoothing_tolerance'] = hypso_smoothing_tolerance
        hypso_increase_factor = wb.cell(row=9, column=10).value
        val_dict['hypso_increase_factor'] = hypso_increase_factor
        hypso_size_max = wb.cell(row=10, column=10).value
        val_dict['hypso_size_max'] = hypso_size_max
        hypso_size_min = wb.cell(row=11, column=10).value
        val_dict['hypso_size_min'] = hypso_size_min

        # Detect conflict
        wb = rule_book['10_DetectConflicts']
        dc_express = wb.cell(row=2, column=2).value
        val_dict['dc_express'] = dc_express
        dc_ref_scale = wb.cell(row=2, column=10).value
        val_dict['dc_ref_scale'] = dc_ref_scale
        dc_severity = wb.cell(row=3, column=10).value
        val_dict['dc_severity'] = dc_severity
        dc_reviewer_session = wb.cell(row=4, column=10).value
        val_dict['dc_reviewer_session'] = dc_reviewer_session
        dc_distance = wb.cell(row=5, column=10).value
        val_dict['dc_distance'] = dc_distance

        # Resolve conflicts buildings
        wb = rule_book['9b_ResolveConflictsBuildings']
        express_val_mx = wb.cell(row=2, column=2).value
        val_dict['express_val_mx'] = express_val_mx
        express_val_mn = wb.cell(row=3, column=2).value
        val_dict['express_val_mn'] = express_val_mn
        visible_field = wb.cell(row=2, column=5).value
        val_dict['visible_field'] = visible_field   
        hierarchy_field = wb.cell(row=244, column=5).value
        val_dict['hierarchy_field'] = hierarchy_field
        search_distance = wb.cell(row=2, column=10).value
        val_dict['search_distance'] = search_distance
        query = wb.cell(row=3, column=3).value
        val_dict['query'] = query   
        bb_lyr_ex = wb.cell(row=120, column=2).value
        val_dict['bb_lyr_ex'] = bb_lyr_ex
        bb_lyr_ex_his = wb.cell(row=228, column=2).value
        val_dict['bb_lyr_ex_his'] = bb_lyr_ex_his
        ref_scale = wb.cell(row=79, column=10).value
        val_dict['ref_scale'] = ref_scale
        minimum_size = wb.cell(row=78, column=10).value
        val_dict['minimum_size'] = minimum_size
        bld_gap = wb.cell(row=77, column=10).value
        val_dict['bld_gap'] = bld_gap
        ap_src_dis_mx = wb.cell(row=245, column=10).value
        val_dict['ap_src_dis_mx'] = ap_src_dis_mx
        ap_src_dis_mn = wb.cell(row=244, column=10).value
        val_dict['ap_src_dis_mn'] = ap_src_dis_mn
        orient_dir = wb.cell(row=246, column=10).value
        val_dict['orient_dir'] = orient_dir
        in_prim_sql = wb.cell(row=311, column=3).value
        val_dict['in_prim_sql'] = in_prim_sql
        max_gap_area = wb.cell(row=311, column=10).value
        val_dict['max_gap_area'] = max_gap_area
        fill_option = wb.cell(row=312, column=10).value
        val_dict['fill_option'] = fill_option

        # Resolve conflicts lines
        wb = rule_book['9a_ResolveConflictsLines']
        ln_lyr_ex = wb.cell(row=2, column=2).value
        val_dict['ln_lyr_ex'] = ln_lyr_ex
        res_con_line_delete = wb.cell(row=2, column=10).value
        val_dict['res_con_line_delete'] = res_con_line_delete
        river_ex = wb.cell(row=5, column=2).value
        val_dict['river_ex'] = river_ex
        road_query_rlc = wb.cell(row=2, column=3).value
        val_dict['road_query_rlc'] = road_query_rlc
        name_fld = wb.cell(row=25, column=5).value
        val_dict['name_fld'] = name_fld
        distance_b = wb.cell(row=26, column=10).value
        val_dict['distance_b'] = distance_b
        distance_s = wb.cell(row=27, column=10).value
        val_dict['distance_s'] = distance_s
        min_area = wb.cell(row=25, column=10).value
        val_dict['min_area'] = min_area
        additional_criteria = wb.cell(row=77, column=10).value
        val_dict['additional_criteria'] = additional_criteria
        distance_rcl = wb.cell(row=61, column=10).value
        val_dict['distance_rcl'] = distance_rcl
        min_length = wb.cell(row=60, column=10).value
        val_dict['min_length'] = min_length
        orient_fld = wb.cell(row=68, column=5).value
        val_dict['orient_fld'] = orient_fld
        offset_dist_l = wb.cell(row=69, column=10).value
        val_dict['offset_dist_l'] = offset_dist_l
        offset_dist_u = wb.cell(row=70, column=10).value
        val_dict['offset_dist_u'] = offset_dist_u
        offset_dist_benc_l = wb.cell(row=74, column=10).value
        val_dict['offset_dist_benc_l'] = offset_dist_benc_l
        offset_dist_benc_u = wb.cell(row=75, column=10).value
        val_dict['offset_dist_benc_u'] = offset_dist_benc_u
        perpendicular_k = wb.cell(row=68, column=10).value
        val_dict['perpendicular_k'] = perpendicular_k
        perpendicular_b = wb.cell(row=73, column=10).value
        val_dict['perpendicular_b'] = perpendicular_b
        bench_query = wb.cell(row=73, column=3).value
        val_dict['bench_query'] = bench_query
        res_con_line_erase_input_fcs = wb.cell(row=65, column=10).value
        val_dict['res_con_line_erase_input_fcs'] = res_con_line_erase_input_fcs

        # Apply carto symbology
        wb = rule_book['8_ApplyCartoSymbology']
        query = wb.cell(row=36, column=3).value
        val_dict['query_acs'] = query
        distance_acs = wb.cell(row=39, column=10).value
        val_dict['distance_acs'] = distance_acs
        mx_no_close_fcs_l = wb.cell(row=37, column=10).value
        val_dict['mx_no_close_fcs_l'] = mx_no_close_fcs_l
        mx_no_close_fcs_m = wb.cell(row=38, column=10).value
        val_dict['mx_no_close_fcs_m'] = mx_no_close_fcs_m
        mx_no_close_fcs_u = wb.cell(row=36, column=10).value
        val_dict['mx_no_close_fcs_u'] = mx_no_close_fcs_u
        feature_count = wb.cell(row=49, column=10).value
        val_dict['feature_count_acs'] = feature_count
        specification = wb.cell(row=50, column=10).value
        val_dict['specification'] = specification

        # Load Data Final 100K
        wb = rule_book['11_LoadDataFinal100K']
        versions = wb.cell(row=2, column=10).value
        val_dict['versions'] = versions
        return val_dict


